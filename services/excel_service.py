# -*- coding: utf-8 -*-
"""
Excel 数据访问层（基于 pandas）。

设计意图：
- 所有与 Excel 的读写集中在此模块，路由层只调用这里的方法。
- 后续若接入 MySQL / PostgreSQL，可新增 repository 模块并实现相同语义接口，
  再让路由依赖抽象接口（或依赖注入），本文件可逐步废弃或改为导出到 DB 的迁移脚本。

总表列名兼容：
- 学号：学号 / StudentID / student_id
- 姓名：姓名 / 名字 / Name
- 时长：总时长 / 时长 / 志愿服务时长 / hours
"""
from __future__ import annotations

import os
import re
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import re

import pandas as pd


# 列名别名 -> 标准列名（内部统一用标准名处理）
_COL_ALIASES = {
    "学号": ["学号", "StudentID", "student_id", "StudentId", "学籍号", "学号ID", "NO", "No", "Number"],
    "姓名": ["姓名", "名字", "Name", "name", "学生姓名", "姓名Name"],
    "总时长": ["总时长", "时长", "志愿服务时长", "hours", "Hours", "志愿时长", "累计时长", "总计时长"],
    "项数": ["项数", "次数", "条数", "Items", "ItemCount", "活动项数", "总项数"],
}

# 录入时在总表上找「最后一列有数据」的扫描宽度上限。部分 xlsx 的 max_column 会被撑到上万列
#（如误选整行格式、残留引用），若按 max_column×max_row 全表扫描会卡死进程与事件循环。
_APPEND_DURATION_MAX_COL_SCAN = 500


def _find_last_column_with_data(sheet) -> int:
    """在不超过 _APPEND_DURATION_MAX_COL_SCAN 的范围内找最右一列存在非空单元格的列号。"""
    scan_cols = min(sheet.max_column, _APPEND_DURATION_MAX_COL_SCAN)
    last_col_idx = 0
    max_row = sheet.max_row
    for c in range(1, scan_cols + 1):
        for r in range(1, max_row + 1):
            val = sheet.cell(row=r, column=c).value
            if val is not None and str(val).strip() != "":
                last_col_idx = c
                break
    return last_col_idx


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """将 DataFrame 列名映射到标准列名（学号、姓名、总时长）；未匹配列保留原样。"""
    rename_map = {}
    
    # 优先进行精确匹配，防止“校外志愿时长”误匹配为“总时长”
    for col in df.columns:
        if pd.isna(col): continue
        clean_col = str(col).strip().replace(" ", "").replace("\n", "").replace("\r", "").lower()
        
        for standard, aliases in _COL_ALIASES.items():
            if standard in rename_map.values(): continue
            # 精确匹配别名
            if any(clean_col == str(alias).lower() for alias in aliases):
                rename_map[col] = standard
                break
                
    # 如果还有没匹配上的标准列，再尝试模糊匹配（包含关系）
    for col in df.columns:
        if col in rename_map or pd.isna(col): continue
        clean_col = str(col).strip().replace(" ", "").replace("\n", "").replace("\r", "").lower()
        
        for standard, aliases in _COL_ALIASES.items():
            if standard in rename_map.values(): continue
            # 模糊匹配：列名中包含别名，且别名长度大于 1（防止误匹配单字）
            # 特殊处理：如果是 2024/2025 级的“班级”列，不应该被识别为“姓名”
            if standard == "姓名" and "班级" in clean_col:
                continue
            if any(len(str(alias)) > 1 and str(alias).lower() in clean_col for alias in aliases):
                rename_map[col] = standard
                break
                
    return df.rename(columns=rename_map)


def _ensure_data_dir(path: str) -> None:
    d = os.path.dirname(path)
    if d and not os.path.isdir(d):
        os.makedirs(d, exist_ok=True)


def load_master_dataframe(master_path: str) -> pd.DataFrame:
    """
    读取总表为 DataFrame；优先读取 CSV，不存在则读取 Excel。
    """
    csv_path = master_path.rsplit(".", 1)[0] + ".csv"
    
    # 优先尝试读取 CSV
    if os.path.isfile(csv_path):
        try:
            df = pd.read_csv(csv_path)
            # CSV 读取时可能把学号存为 float，需要预处理
            if "学号" in df.columns:
                df["学号"] = df["学号"].astype(str).str.replace(r"\.0$", "", regex=True)
            return _normalize_columns(df)
        except Exception as e:
            print(f"Error reading CSV {csv_path}: {e}")

    if not os.path.isfile(master_path):
        return pd.DataFrame(columns=["学号", "姓名", "总时长"])

    # engine 由 openpyxl 处理 xlsx；xls 需 xlrd，此处仅开放 xlsx 上传为主流
    try:
        df = pd.read_excel(master_path)
        df = _normalize_columns(df)
        return df
    except Exception:
        return pd.DataFrame(columns=["学号", "姓名", "总时长"])


def save_master_as_csv(master_path: str) -> bool:
    """
    将 Excel 总表另存为一份 CSV 文件，以便快速查询。
    CSV 会固定所有公式为数值。
    """
    try:
        if not os.path.isfile(master_path):
            return False
        
        # pandas 读取 Excel 时会自动获取公式计算后的值
        df = pd.read_excel(master_path)
        csv_path = master_path.rsplit(".", 1)[0] + ".csv"
        
        # 保存为 CSV，使用 utf-8-sig 编码以兼容 Excel 打开中文
        df.to_csv(csv_path, index=False, encoding="utf-8-sig")
        return True
    except Exception as e:
        print(f"Error converting Excel to CSV: {e}")
        return False


def _process_excel_bytes_to_dfs(file_bytes: bytes) -> List[pd.DataFrame]:
    """
    将上传的 Excel 二进制内容解析为多个 DataFrame（每个分表一个）。
    """
    from io import BytesIO
    xl = pd.ExcelFile(BytesIO(file_bytes))
    sheet_names = xl.sheet_names
    dfs = []
    
    for sheet_name in sheet_names:
        # 1. 先尝试正常读取（ header=0 ）
        try:
            df = pd.read_excel(xl, sheet_name=sheet_name)
            df = _normalize_columns(df)
            
            # 如果没找到姓名或学号，尝试向下探测表头（ header=None ）
            if "姓名" not in df.columns or "学号" not in df.columns:
                df_raw = pd.read_excel(xl, sheet_name=sheet_name, header=None)
                header_idx = -1
                for i in range(min(15, len(df_raw))):
                    row_str = "".join([str(x) for x in df_raw.iloc[i].dropna()]).replace(" ", "").lower()
                    if any(kw.lower() in row_str for kw in ["姓名", "学号", "Name", "Student"]):
                        header_idx = i
                        break
                
                if header_idx != -1:
                    df = df_raw.iloc[header_idx:].copy()
                    df.columns = df_raw.iloc[header_idx]
                    df = df.iloc[1:].reset_index(drop=True)
                    df = _normalize_columns(df)
            
            # 过滤掉全为空的行，且至少包含姓名或学号
            if not df.empty and ("姓名" in df.columns or "学号" in df.columns):
                df = df.dropna(how='all')
                if not df.empty:
                    dfs.append(df)
        except Exception as e:
            print(f"Error processing sheet {sheet_name}: {e}")
            continue
    return dfs


def merge_and_save_master(file_bytes: bytes, dest_path: str) -> bool:
    """
    将上传的 Excel 文件中的所有分表合并为一张表，并保存到指定路径。
    """
    try:
        dfs = _process_excel_bytes_to_dfs(file_bytes)
        
        if not dfs:
            from io import BytesIO
            # 如果实在探测不到有效数据，退回最原始的暴力保存模式
            df_fallback = pd.read_excel(BytesIO(file_bytes))
            os.makedirs(os.path.dirname(dest_path), exist_ok=True)
            df_fallback.to_excel(dest_path, index=False, engine="openpyxl")
            return True
            
        # 合并所有分表
        merged_df = pd.concat(dfs, ignore_index=True)
        os.makedirs(os.path.dirname(dest_path), exist_ok=True)
        merged_df.to_excel(dest_path, index=False, engine="openpyxl")
        return True
    except Exception as e:
        print(f"Error merging sheets: {e}")
        from io import BytesIO
        # 最后一道防线
        try:
            df_final = pd.read_excel(BytesIO(file_bytes))
            df_final.to_excel(dest_path, index=False, engine="openpyxl")
            return True
        except:
            return False


def merge_multiple_excels_to_bytes(files_data: List[bytes]) -> Optional[bytes]:
    """
    将多个 Excel 文件中的所有分表合并为一张表，并返回 Excel 二进制内容。
    """
    try:
        all_dfs = []
        for file_bytes in files_data:
            dfs = _process_excel_bytes_to_dfs(file_bytes)
            all_dfs.extend(dfs)
            
        if not all_dfs:
            return None
            
        merged_df = pd.concat(all_dfs, ignore_index=True)
        
        from io import BytesIO
        buf = BytesIO()
        merged_df.to_excel(buf, index=False, engine="openpyxl")
        buf.seek(0)
        return buf.read()
    except Exception as e:
        print(f"Error merging multiple excels: {e}")
        return None


def append_durations_to_master(
    pending_bytes: bytes, 
    master_bytes: bytes, 
    activity_name: str,
    multi_sheet: bool = False
) -> Tuple[Optional[bytes], List[str], int, Dict[str, int]]:
    """
    将待录入时长表中的时长追加到原时长总表的新列中，同时保留原表公式。
    
    :param pending_bytes: 待录入时长表文件二进制
    :param master_bytes: 原时长总表文件二进制
    :param activity_name: 活动名称（即待录入表的文件名，用作新列表头）
    :param multi_sheet: 是否在所有包含姓名列的工作表中执行录入（针对22、23级）
    :return: (更新后的 Excel 二进制, 未匹配到的姓名列表，成功匹配的数量，按年级成功匹配的数量)
    """
    from io import BytesIO
    import openpyxl
    from copy import copy
    try:
        # 1. 读取待录入表 (File A) - 关键：使用 data_only=True 以获取公式计算后的数字结果
        wb_pending = openpyxl.load_workbook(BytesIO(pending_bytes), data_only=True)
        # 显式读取第一张工作表，忽略其他 Sheet
        ws_pending = wb_pending.worksheets[0]
        
        # 寻找待录入表的“姓名”、“时长”、“年级”列
        pending_name_col = -1
        pending_time_col = -1
        pending_grade_col = -1
        pending_clazz_col = -1
        pending_header_row = -1
        
        # 探测前 20 行
        for r in range(1, min(21, ws_pending.max_row + 1)):
            for cell in ws_pending[r]:
                if cell.value is not None:
                    v = str(cell.value).strip().replace(" ", "").lower()
                    if v in ["姓名", "名字", "name"]:
                        pending_name_col = cell.column
                        pending_header_row = r
                    elif v in ["时长", "总时长", "hours"]:
                        pending_time_col = cell.column
                    elif "班级" in v or v in ["class", "clazz", "classid"]:
                        pending_clazz_col = cell.column
                    # 年级列表头可能不是固定的“年级”，而是类似“2024级本科生”“25级”等
                    # 这里做模糊识别：包含“年级/grade”关键词，或包含“数字+级”模式（排除“班级”）
                    elif (
                        "年级" in v
                        or "grade" in v
                        or (
                            ("级" in v)
                            and ("班级" not in v)
                            and re.search(r"\d{2,4}\s*级", v) is not None
                        )
                    ):
                        pending_grade_col = cell.column
            if pending_name_col != -1 and pending_time_col != -1:
                break
        
        if pending_name_col == -1 or pending_time_col == -1:
            return None, [], 0, {}

        pending_data = {}
        # 从表头下一行开始读取数据
        for r in range(pending_header_row + 1, ws_pending.max_row + 1):
            name_val = ws_pending.cell(row=r, column=pending_name_col).value
            time_val = ws_pending.cell(row=r, column=pending_time_col).value
            grade_val = ws_pending.cell(row=r, column=pending_grade_col).value if pending_grade_col != -1 else None
            clazz_val = ws_pending.cell(row=r, column=pending_clazz_col).value if pending_grade_col == -1 and pending_clazz_col != -1 else None
            
            if name_val is not None:
                name_str = str(name_val).strip()
                if name_str:
                    # 规则：如果时长格内容是“见xx”形式，则跳过该条记录
                    if isinstance(time_val, str) and time_val.strip().startswith("见"):
                        continue

                    try:
                        time_num = float(time_val) if time_val is not None else 0.0
                    except (ValueError, TypeError):
                        continue # 如果转换失败，跳过
                    
                    # 规则：待录入时长表如果有 0 则不录入
                    if time_num <= 0:
                        continue

                    grade_str = None
                    if grade_val:
                        grade_str = str(grade_val).strip()
                    elif clazz_val:
                        # 从“班级”内容里抽取类似“2024级”
                        m = re.search(r"(\d{2,4}\s*级)", str(clazz_val))
                        grade_str = m.group(1) if m else None
                    # 统一年级展示粒度：只保留类似 “2025级” 的部分
                    if grade_str:
                        m2 = re.search(r"(\d{2,4})\s*级", grade_str)
                        grade_str = (m2.group(1) + "级") if m2 else grade_str
                    grade_str = grade_str if grade_str else "未知年级"
                    pending_data[name_str] = {"val": time_num, "grade": grade_str}
        
        if not pending_data:
            return None, [], 0, {}

        # 2. 使用 openpyxl 加载原时长总表 (File B) - 关键：data_only=False 以保留公式
        wb_master = openpyxl.load_workbook(BytesIO(master_bytes), data_only=False)
        
        def copy_style(src_cell, tgt_cell):
            """全面复制单元格样式"""
            if src_cell.has_style:
                tgt_cell.font = copy(src_cell.font)
                tgt_cell.border = copy(src_cell.border)
                tgt_cell.fill = copy(src_cell.fill)
                tgt_cell.number_format = copy(src_cell.number_format)
                tgt_cell.protection = copy(src_cell.protection)
                tgt_cell.alignment = copy(src_cell.alignment)

        matched_names = set()
        success_count = 0
        grade_success_counts: Dict[str, int] = {}
        name_aliases = ["姓名", "名字", "name", "学生姓名"]
        sheets_processed = 0

        # 3. 遍历工作表执行录入
        for sheet in wb_master.worksheets:
            master_name_col = -1
            master_header_row = -1
            
            # 探测表头
            for r in range(1, min(31, sheet.max_row + 1)):
                for cell in sheet[r]:
                    if cell.value is not None:
                        v = str(cell.value).strip().replace(" ", "").lower()
                        if any(alias in v for alias in name_aliases):
                            master_name_col = cell.column
                            master_header_row = r
                            break
                if master_name_col != -1: break
            
            if master_name_col == -1: continue

            # 如果不是 multi_sheet 模式且已经处理过一个工作表，则跳过后续
            if not multi_sheet and sheets_processed >= 1: break

            # 收集该 Sheet 中存在的姓名和对应行号
            sheet_matches = {} # {row_idx: pending_key}
            for r in range(master_header_row + 1, sheet.max_row + 1):
                m_name_val = sheet.cell(row=r, column=master_name_col).value
                if m_name_val is not None:
                    m_name = str(m_name_val).strip()
                    if m_name in pending_data:
                        sheet_matches[r] = m_name

            # 只有当该 Sheet 中有匹配项时才新增列
            if not sheet_matches: continue

            # 确定新列位置（寻找有任何内容的最后一列，不仅仅是表头行）
            last_col_idx = _find_last_column_with_data(sheet)

            new_col_idx = last_col_idx + 1
            header_cell = sheet.cell(row=master_header_row, column=new_col_idx, value=activity_name)
            
            # 4.1 复制格式逻辑
            from copy import copy
            import re as re_mod
            def copy_style(src_cell, tgt_cell):
                """全面复制单元格样式"""
                if src_cell.has_style:
                    tgt_cell.font = copy(src_cell.font)
                    tgt_cell.border = copy(src_cell.border)
                    tgt_cell.fill = copy(src_cell.fill)
                    tgt_cell.number_format = copy(src_cell.number_format)
                    tgt_cell.protection = copy(src_cell.protection)
                    tgt_cell.alignment = copy(src_cell.alignment)

            # 确定样式源列 (style_source_col_idx)
            style_source_col_idx = last_col_idx
            if style_source_col_idx > 0:
                # 如果前一列表头是 "24-25学年" 这种格式，则向前回溯一列作为样式源
                prev_header_val = str(sheet.cell(row=master_header_row, column=style_source_col_idx).value or "").strip()
                # 匹配类似 "24-25学年", "2024-2025学年" 等
                if re_mod.search(r'\d+-\d+学年', prev_header_val) and style_source_col_idx > 1:
                    style_source_col_idx -= 1

            if style_source_col_idx > 0:
                # 复制表头格式
                src_header_cell = sheet.cell(row=master_header_row, column=style_source_col_idx)
                copy_style(src_header_cell, header_cell)
                # 复制列宽
                src_col_letter = openpyxl.utils.get_column_letter(style_source_col_idx)
                new_col_letter = openpyxl.utils.get_column_letter(new_col_idx)
                if src_col_letter in sheet.column_dimensions:
                    sheet.column_dimensions[new_col_letter].width = sheet.column_dimensions[src_col_letter].width

            # 写入匹配到的数据并复制行样式
            for r in range(master_header_row + 1, sheet.max_row + 1):
                target_cell = sheet.cell(row=r, column=new_col_idx)
                if style_source_col_idx > 0:
                    copy_style(sheet.cell(row=r, column=style_source_col_idx), target_cell)
                
                if r in sheet_matches:
                    name_key = sheet_matches[r]
                    target_cell.value = pending_data[name_key]["val"]
                    matched_names.add(name_key)
                    success_count += 1
                    grade_str = pending_data[name_key].get("grade") or "未知年级"
                    grade_success_counts[grade_str] = grade_success_counts.get(grade_str, 0) + 1
            
            sheets_processed += 1

        if sheets_processed == 0:
            return None, [], 0, {}

        # 6. 收集未匹配名单
        failed_names = []
        for name, info in pending_data.items():
            if name not in matched_names:
                failed_names.append(f"{name}({info['grade']})")
        
        # 7. 保存并返回
        out_buf = BytesIO()
        wb_master.save(out_buf)
        out_buf.seek(0)
        
        return out_buf.read(), failed_names, success_count, grade_success_counts

    except Exception as e:
        print(f"Error in append_durations_to_master (openpyxl v6): {e}")
        return None, [], 0, {}


def aggregate_student_stats(df: pd.DataFrame) -> pd.DataFrame:
    """
    按学号聚合：数值列求和，非数值列保留首个。
    """
    if df.empty:
        return pd.DataFrame(columns=["学号", "姓名", "总时长", "项数"])

    work = df.copy()
    if "总时长" not in work.columns:
        work["总时长"] = 0
    work["总时长"] = pd.to_numeric(work["总时长"], errors="coerce").fillna(0)
    
    if "学号" in work.columns:
        work["学号"] = _normalize_id(work["学号"])
    
    if "项数" in work.columns:
        work["项数"] = pd.to_numeric(work["项数"], errors="coerce").fillna(0)

    # 重新计算项数（统计“项数”列之后的所有数值格）
    if "项数" in work.columns:
        try:
            item_idx = list(work.columns).index("项数")
            activity_cols = list(work.columns)[item_idx + 1 :]
            if activity_cols:
                numeric_activity = work[activity_cols].apply(lambda col: pd.to_numeric(col, errors="coerce"))
                work["项数"] = numeric_activity.notna().sum(axis=1)
        except Exception:
            pass

    # 找出所有列，准备聚合
    all_cols = list(work.columns)
    if "学号" in all_cols and work["学号"].notna().any():
        agg_dict = {}
        for col in all_cols:
            if col == "学号": continue
            # 尝试转换为数值，如果成功则求和，否则取第一个
            try:
                pd.to_numeric(work[col], errors="raise")
                agg_dict[col] = "sum"
            except (ValueError, TypeError):
                agg_dict[col] = "first"
        
        agg = work.groupby("学号", dropna=False).agg(agg_dict).reset_index()
        return agg
    else:
        # 如果没有学号，按姓名聚合
        if "姓名" in all_cols:
            agg_dict = {}
            for col in all_cols:
                if col == "姓名": continue
                try:
                    pd.to_numeric(work[col], errors="raise")
                    agg_dict[col] = "sum"
                except (ValueError, TypeError):
                    agg_dict[col] = "first"
            agg = work.groupby("姓名", dropna=False).agg(agg_dict).reset_index()
            return agg
    
    return work


def search_student(
    master_path: str, keyword: str
) -> Tuple[bool, str, List[Dict[str, Any]]]:
    """
    按姓名或学号模糊查询（keyword 同时匹配两列）。

    返回：(是否成功, 错误信息或空字符串, 结果列表)
    """
    keyword = (keyword or "").strip()
    if not keyword:
        return False, "请输入姓名或学号", []

    df = load_master_dataframe(master_path)
    if df.empty:
        return False, "暂无志愿时长总表数据，请等待管理员上传", []

    agg = aggregate_student_stats(df)
    kw = keyword
    mask = pd.Series([False] * len(agg))
    if "学号" in agg.columns:
        mask = mask | agg["学号"].astype(str).str.contains(kw, case=False, na=False)
    if "姓名" in agg.columns:
        mask = mask | agg["姓名"].astype(str).str.contains(kw, case=False, na=False)

    hit = agg[mask]
    records: List[Dict[str, Any]] = []
    for _, row in hit.iterrows():
        records.append(
            {
                "学号": str(row.get("学号", "") or ""),
                "姓名": str(row.get("姓名", "") or ""),
                "总时长": float(row.get("总时长", 0) or 0),
                "项数": int(row.get("项数", 0) or 0),
            }
        )
    return True, "", records


def _normalize_id(series: pd.Series) -> pd.Series:
    """统一学号格式：转为字符串，去掉 .0，去空格。"""
    return series.astype(str).str.strip().str.replace(r"\.0$", "", regex=True).replace(["nan", "None", "NAN", "null"], "")


def search_student_by_name_and_id(
    master_path: str, *, name: str, student_id: str
) -> Tuple[bool, str, List[Dict[str, Any]]]:
    """
    通过“姓名 + 学号”查询同一学号下匹配的记录并聚合。
    """
    # 去掉用户输入中的空格
    name = (name or "").strip()
    student_id = (student_id or "").strip()
    if not name or not student_id:
        return False, "请输入姓名与学号", []

    df = load_master_dataframe(master_path)
    if df.empty:
        return False, "暂无志愿时长总表数据，请等待管理员上传", []

    if "学号" not in df.columns or "姓名" not in df.columns:
        return False, "总表缺少“学号”或“姓名”列，请检查上传的 Excel 格式", []

    work = df.copy()
    # 关键修复：统一学号格式后再匹配
    work["学号"] = _normalize_id(work["学号"])
    
    # 姓名做空白归一化
    work["姓名"] = work["姓名"].astype(str).str.replace(r"\s+", "", regex=True)
    name_norm = re.sub(r"\s+", "", str(name))

    # 使用处理后的学号进行匹配
    mask_sid = work["学号"] == student_id
    hit = work[mask_sid]
    
    # 如果学号匹配到了，再校验姓名是否匹配（忽略空格）
    if not hit.empty:
        # 只要有一行姓名匹配即可（考虑到总表合并后可能有多行）
        name_match = hit["姓名"].eq(name_norm).any()
        if not name_match:
            return True, "学号正确但姓名不匹配，请检查输入", []
    else:
        return True, "未找到该学号对应的记录", []
    
    # 聚合结果
    agg = aggregate_student_stats(hit)
    records: List[Dict[str, Any]] = []
    
    # 提取所有活动详情（排除标准列）
    # 使用 set 进行快速查找，并包含所有可能出现的非活动列名
    exclude_cols = {"学号", "姓名", "总时长", "项数", "序号", "专业", "班级", "志愿活动所处学年", "活动名称", "时长", "活动项数", "微信号联系方式", "处理状态", "管理员备注"}
    
    academic_year_pattern = re.compile(r"^\d{2,4}\s*-\s*\d{2,4}\s*学年$")

    for _, row in agg.iterrows():
        record = {
            "学号": str(row.get("学号", "") or ""),
            "姓名": str(row.get("姓名", "") or ""),
            "总时长": float(row.get("总时长", 0) or 0),
            "项数": int(row.get("项数", 0) or 0),
            "活动列表": [] # 存储结构化项：header/activity
        }
        # 遍历该行的所有列
        for col in agg.columns:
            col_str = str(col).strip()
            clean_col = col_str.replace(" ", "").replace("\n", "").replace("\r", "")
            
            # 如果不是排除列
            if clean_col not in exclude_cols:
                val = row[col]
                if pd.notna(val):
                    # 学年列作为分组标题展示，而不是活动项
                    if academic_year_pattern.match(clean_col):
                        record["活动列表"].append({"type": "header", "title": col_str})
                        continue

                    try:
                        num_val = float(val)
                        if num_val > 0:
                            # 记录活动项与数值
                            record["活动列表"].append({"type": "activity", "name": col_str, "value": num_val})
                    except (ValueError, TypeError):
                        s_val = str(val).strip()
                        if s_val and s_val not in ["0", "0.0", "nan", "None"]:
                            # 避免把“24-25学年: 24-25学年”这类重复值当活动项
                            if clean_col == s_val.replace(" ", "").replace("\n", "").replace("\r", ""):
                                continue
                            record["活动列表"].append({"type": "activity", "name": col_str, "value": s_val})
        records.append(record)
    return True, "", records


def search_student_multi_sheets(
    master_path: str, *, name: str, student_id: str
) -> Tuple[bool, str, List[Dict[str, Any]]]:
    """
    针对多分表的 Excel 进行查询。
    """
    sheets = load_all_sheets(master_path)
    if not sheets:
        return False, "暂无志愿时长总表数据，请等待管理员上传", []

    all_records: List[Dict[str, Any]] = []
    
    # 归一化用户输入的学号和姓名
    name = (name or "").strip()
    student_id = (student_id or "").strip()
    name_norm = re.sub(r"\s+", "", str(name))

    for sheet_name, df in sheets:
        if "学号" not in df.columns or "姓名" not in df.columns:
            continue
            
        work = df.copy()
        work["学号"] = work["学号"].astype(str).str.strip()
        work["姓名"] = work["姓名"].astype(str).str.replace(r"\s+", "", regex=True)

        mask_sid = work["学号"].eq(student_id)
        hit = work[mask_sid]
        if hit.empty:
            continue

        agg = aggregate_student_stats(hit)
        for _, row in agg.iterrows():
            all_records.append(
                {
                    "学号": str(row.get("学号", "") or ""),
                    "姓名": str(row.get("姓名", "") or ""),
                    "总时长": float(row.get("总时长", 0) or 0),
                    "项数": int(row.get("项数", 0) or 0),
                    "分表名": sheet_name, # 记录来源分表
                }
            )
    
    return True, "", all_records


FEEDBACK_COLUMNS = [
    "提交时间",
    "姓名",
    "班级",
    "年级",
    "问题说明",
    "证明文件相对路径",
    "处理状态",
]


def append_feedback_row(
    feedback_path: str,
    *,
    name: str,
    clazz: str,
    grade: str,
    note: str,
    proof_relative_path: str,
) -> None:
    """
    将一条反馈追加写入反馈汇总 Excel。

    后续接入数据库时，可改为 INSERT INTO feedback (...) VALUES (...)。
    """
    _ensure_data_dir(feedback_path)
    row = {
        "提交时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "姓名": name,
        "班级": clazz,
        "年级": grade,
        "问题说明": note,
        "证明文件相对路径": proof_relative_path,
        "处理状态": "待处理",
    }

    if os.path.isfile(feedback_path):
        old = pd.read_excel(feedback_path)
        for c in FEEDBACK_COLUMNS:
            if c not in old.columns:
                old[c] = ""
        new = pd.concat([old, pd.DataFrame([row])], ignore_index=True)
    else:
        new = pd.DataFrame([row], columns=FEEDBACK_COLUMNS)

    new.to_excel(feedback_path, index=False, engine="openpyxl")


def load_feedback_dataframe(feedback_path: str) -> pd.DataFrame:
    if not os.path.isfile(feedback_path):
        return pd.DataFrame(columns=FEEDBACK_COLUMNS)
    df = pd.read_excel(feedback_path)
    for c in FEEDBACK_COLUMNS:
        if c not in df.columns:
            df[c] = ""
    return df


def export_feedback_summary_bytes(feedback_path: str) -> Optional[bytes]:
    """读取反馈表并返回 xlsx 二进制内容，供下载；无数据时返回 None。"""
    df = load_feedback_dataframe(feedback_path)
    if df.empty:
        return None
    from io import BytesIO

    buf = BytesIO()
    df.to_excel(buf, index=False, engine="openpyxl")
    buf.seek(0)
    return buf.read()


def delete_topup_row(
    topup_path: str,
    *,
    submit_time: str,
    name: str,
) -> bool:
    """
    根据“提交时间”和“姓名”删除一条补录记录。
    """
    if not os.path.isfile(topup_path):
        return False

    df = pd.read_excel(topup_path)
    # 确保列名一致且为 object 类型，防止 pandas 自动识别为 float 导致无法写入字符串
    for c in TOPUP_COLUMNS:
        if c not in df.columns:
            df[c] = ""
        df[c] = df[c].astype(object)

    # 匹配提交时间和姓名（trim 后比较）
    mask = (df["提交时间"].astype(str).str.strip() == submit_time.strip()) & \
           (df["姓名"].astype(str).str.strip() == name.strip())

    if not mask.any():
        return False

    # 删除匹配的行（仅删除第一条匹配项，防止误删多条）
    idx_to_drop = df[mask].index[0]
    new_df = df.drop(index=idx_to_drop)

    new_df.to_excel(topup_path, index=False, engine="openpyxl")
    return True


def delete_topup_rows_batch(
    topup_path: str,
    items: list[dict[str, str]],
) -> int:
    """
    批量删除补录记录。
    items: [{"submit_time": "...", "name": "..."}, ...]
    返回成功删除的条数。
    """
    if not os.path.isfile(topup_path) or not items:
        return 0

    df = pd.read_excel(topup_path)
    # 确保列名一致且为 object 类型
    for c in TOPUP_COLUMNS:
        if c not in df.columns:
            df[c] = ""
        df[c] = df[c].astype(object)

    initial_count = len(df)
    
    # 构建过滤条件
    for item in items:
        s_time = str(item.get("submit_time", "")).strip()
        s_name = str(item.get("name", "")).strip()
        
        # 匹配提交时间和姓名
        mask = (df["提交时间"].astype(str).str.strip() == s_time) & \
               (df["姓名"].astype(str).str.strip() == s_name)
        
        if mask.any():
            # 删除匹配的所有行（理论上同一时间同一人只有一条，但这里支持删除所有匹配项）
            df = df[~mask]

    final_count = len(df)
    deleted_count = initial_count - final_count

    if deleted_count > 0:
        df.to_excel(topup_path, index=False, engine="openpyxl")

    return deleted_count


def update_topup_rows_batch(
    topup_path: str,
    items: list[dict[str, str]],
    *,
    status: str = None,
    admin_note: str = None,
) -> int:
    """
    批量更新补录记录的状态或管理员备注。
    items: [{"submit_time": "...", "name": "..."}, ...]
    """
    if not os.path.isfile(topup_path) or not items:
        return 0

    df = pd.read_excel(topup_path)
    # 确保列名一致且为 object 类型
    for c in TOPUP_COLUMNS:
        if c not in df.columns:
            df[c] = ""
        df[c] = df[c].astype(object)

    updated_count = 0
    for item in items:
        s_time = str(item.get("submit_time", "")).strip()
        s_name = str(item.get("name", "")).strip()
        
        mask = (df["提交时间"].astype(str).str.strip() == s_time) & \
               (df["姓名"].astype(str).str.strip() == s_name)
        
        if mask.any():
            if status is not None:
                df.loc[mask, "处理状态"] = status
            if admin_note is not None:
                df.loc[mask, "管理员备注"] = admin_note
            updated_count += mask.sum()

    if updated_count > 0:
        df.to_excel(topup_path, index=False, engine="openpyxl")

    return updated_count


# ---------------------------------------------------------------------------
# 时长补录：校内 / 校外
# ---------------------------------------------------------------------------

TOPUP_COLUMNS = [
    "提交时间",
    "补录类型",  # 校内/校外
    "姓名",
    "班级",
    "年级",
    "志愿活动所处学年",
    "活动名称",
    "时长",
    "活动项数",  # 校外必填；校内为空
    "活动描述",  # 校外必填；校内为空
    "证明文件相对路径",
    "本人活动照片相对路径",  # 校外必填；校内为空
    "微信号联系方式",
    "备注",
    "处理状态",
    "管理员备注",
]


def append_topup_row(
    topup_path: str,
    *,
    topup_type: str,  # 校内/校外
    name: str,
    clazz: str,
    grade: str,
    academic_year: str,
    activity_name: str,
    duration: int,
    item_count: int | None = None,
    description: str = "",
    proof_relative_path: str,
    photo_relative_path: str = "",
    wechat: str = "",
    note: str = "",
) -> None:
    """
    将一条“时长补录”追加写入补录汇总 Excel。
    """
    _ensure_data_dir(topup_path)
    row = {
        "提交时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "补录类型": topup_type,
        "姓名": name,
        "班级": clazz,
        "年级": grade,
        "志愿活动所处学年": academic_year,
        "活动名称": activity_name,
        "时长": duration,
        "活动项数": item_count if item_count is not None else "",
        "活动描述": description or "",
        "证明文件相对路径": proof_relative_path,
        "本人活动照片相对路径": photo_relative_path or "",
        "微信号联系方式": wechat or "",
        "备注": note or "",
        "处理状态": "待处理",
    }

    if os.path.isfile(topup_path):
        old = pd.read_excel(topup_path)
        for c in TOPUP_COLUMNS:
            if c not in old.columns:
                old[c] = ""
        new = pd.concat([old, pd.DataFrame([row])], ignore_index=True)
    else:
        new = pd.DataFrame([row], columns=TOPUP_COLUMNS)

    new.to_excel(topup_path, index=False, engine="openpyxl")


def load_topup_dataframe(topup_path: str) -> pd.DataFrame:
    if not os.path.isfile(topup_path):
        return pd.DataFrame(columns=TOPUP_COLUMNS)
    df = pd.read_excel(topup_path)
    for c in TOPUP_COLUMNS:
        if c not in df.columns:
            df[c] = ""
    return df


def export_topup_summary_bytes(topup_path: str) -> Optional[bytes]:
    """读取补录表并返回 xlsx 二进制内容，供下载；无数据时返回 None。"""
    df = load_topup_dataframe(topup_path)
    if df.empty:
        return None

    from io import BytesIO

    buf = BytesIO()
    df.to_excel(buf, index=False, engine="openpyxl")
    buf.seek(0)
    return buf.read()
